import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


EXCEL_FILE = "Predicciones - Porra.xlsx"
OUTPUT_DIR = Path("data")
OUTPUT_FILE = OUTPUT_DIR / "predictions.json"

SHEET_NAME = "Respuestas de formulario 1"

GROUPS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

TEAM_NAME_MAP = {
    "México": "Mexico",
    "Sudáfrica": "South Africa",
    "República de Corea": "South Korea",
    "Chequia": "Czech Republic",
    "Canada": "Canada",
    "Bosnia y Herzegovina": "Bosnia and Herzegovina",
    "Catar": "Qatar",
    "Suiza": "Switzerland",
    "Brasil": "Brazil",
    "Marruecos": "Morocco",
    "Haití": "Haiti",
    "Escocia": "Scotland",
    "EE.UU.": "United States",
    "Paraguay": "Paraguay",
    "Australia": "Australia",
    "Turquía": "Turkey",
    "Alemania": "Germany",
    "Curazao": "Curaçao",
    "Costa de Marfil": "Ivory Coast",
    "Ecuador": "Ecuador",
    "Paises Bajos": "Netherlands",
    "Japón": "Japan",
    "Suecia": "Sweden",
    "Túnez": "Tunisia",
    "Bélgica": "Belgium",
    "Egipto": "Egypt",
    "Irán": "Iran",
    "Nueva Zelanda": "New Zealand",
    "España": "Spain",
    "Cabo Verde": "Cape Verde",
    "Arabia Saudi": "Saudi Arabia",
    "Uruguay": "Uruguay",
    "Francia": "France",
    "Senegal": "Senegal",
    "Irak": "Iraq",
    "Noruega": "Norway",
    "Argentina": "Argentina",
    "Argelia": "Algeria",
    "Austria": "Austria",
    "Jordamia": "Jordan",
    "Jordania": "Jordan",
    "Portugal": "Portugal",
    "RD Congo": "Democratic Republic of the Congo",
    "Uzbekistán": "Uzbekistan",
    "Colombia": "Colombia",
    "Inglaterra": "England",
    "Croacia": "Croatia",
    "Ghana": "Ghana",
    "Panamá": "Panama",
}


def clean_position(value):
    if value is None:
        return None

    value = str(value).strip()
    value = value.replace("º", "").replace("ª", "")

    try:
        return int(value)
    except ValueError:
        return None


def normalize_team_name(team_name):
    if team_name is None:
        return None

    team_name = str(team_name).strip()
    return TEAM_NAME_MAP.get(team_name, team_name)


def read_predictions():
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook[SHEET_NAME]

    participants = []

    # En tu Excel:
    # Fila 1 = nombres de grupos
    # Fila 2 = equipos / preguntas
    # Fila 3 en adelante = respuestas de personas

    for row in range(3, sheet.max_row + 1):
        person_name = sheet.cell(row=row, column=1).value

        if not person_name:
            continue

        person_name = str(person_name).strip()

        person_data = {
            "name": person_name,
            "groups": {},
            "spain_questions": {
                "top_scorer": sheet.cell(row=row, column=50).value,
                "top_assistant": sheet.cell(row=row, column=51).value,
                "goals_for": sheet.cell(row=row, column=52).value,
                "goals_against": sheet.cell(row=row, column=53).value,
            },
        }

        start_col = 2

        for group_index, group_name in enumerate(GROUPS):
            group_start_col = start_col + group_index * 4
            predictions_by_position = {}

            for offset in range(4):
                col = group_start_col + offset

                original_team_name = sheet.cell(row=2, column=col).value
                normalized_team_name = normalize_team_name(original_team_name)

                predicted_position = clean_position(sheet.cell(row=row, column=col).value)

                if predicted_position is None:
                    continue

                predictions_by_position[str(predicted_position)] = {
                    "team": normalized_team_name,
                    "original_team": original_team_name,
                }

            ordered_predictions = []

            for position in range(1, 5):
                position_key = str(position)
                team_prediction = predictions_by_position.get(position_key)

                ordered_predictions.append({
                    "position": position,
                    "team": team_prediction["team"] if team_prediction else None,
                    "original_team": team_prediction["original_team"] if team_prediction else None,
                })

            person_data["groups"][group_name] = ordered_predictions

        participants.append(person_data)

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": EXCEL_FILE,
        "participants": participants,
    }


def save_predictions(data):
    OUTPUT_DIR.mkdir(exist_ok=True)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)

    print(f"Fichero generado: {OUTPUT_FILE}")
    print(f"Participantes encontrados: {len(data['participants'])}")


def main():
    predictions = read_predictions()
    save_predictions(predictions)

    print("\nResumen:")
    for participant in predictions["participants"]:
        print(f"- {participant['name']}")


if __name__ == "__main__":
    main()